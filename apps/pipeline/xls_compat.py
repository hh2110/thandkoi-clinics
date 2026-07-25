"""In-memory conversion of legacy ``.xls`` exports to ``.xlsx``.

The clinic system's real daily export (first sample landed 2026-07-22) is an
OLE2 ``.xls`` (old Excel BIFF format), but the whole pipeline downstream —
``ParserRegistry.sniff_all``, every parser, the upload view's workbook open —
is openpyxl and therefore ``.xlsx``-only. Rather than teach every consumer a
second engine, the upload boundary converts a detected ``.xls`` to an
equivalent in-memory ``.xlsx`` once, and nothing downstream changes.

Privacy invariant #1 holds throughout: the source bytes arrive in memory
(``MemoryFileUploadHandler``), xlrd reads them from memory
(``file_contents=``), and the converted workbook is saved to an in-memory
``BytesIO``. Nothing here touches disk, a model, or a log.
"""

from __future__ import annotations

import io
import zipfile
from typing import BinaryIO

import xlrd
from openpyxl import Workbook

#: First 8 bytes of every OLE2 compound document (the .xls container).
OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

#: Zip-bomb bounds for a directly-uploaded ``.xlsx`` (Plan 15 Track C5). An
#: ``.xlsx`` is a ZIP archive; ``MemoryFileUploadHandler`` already caps the
#: *compressed* upload at ``FILE_UPLOAD_MAX_MEMORY_SIZE`` (2.5 MB), but a
#: crafted archive can still declare gigabytes of uncompressed content — a
#: huge ``sharedStrings.xml`` or sheet part — that ``openpyxl.load_workbook``
#: would materialise into memory and OOM the worker. We bound the archive's
#: *declared* uncompressed size and its compression ratio before opening it,
#: which is what ``load_workbook`` (and every parser) trusts anyway. These are
#: deliberately generous: a real clinic day is well under a megabyte
#: uncompressed, so 100 MB / 200x rejects a bomb without touching any genuine
#: export.
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200


class XlsxTooLargeError(Exception):
    """A ``.xlsx`` whose declared uncompressed size or compression ratio
    exceeds the safe bounds (:data:`MAX_XLSX_UNCOMPRESSED_BYTES` /
    :data:`MAX_XLSX_COMPRESSION_RATIO`).

    Raised by :func:`guard_xlsx_decompression` *before* ``load_workbook``
    reads the archive into memory, so a decompression-bomb upload is rejected
    with the same friendly "nothing was saved" error as any other malformed
    file — never an OOM-killed worker.
    """


def guard_xlsx_decompression(buffer: BinaryIO) -> None:
    """Reject a ``.xlsx`` zip-bomb by its declared metadata; restore position.

    Reads only the ZIP central directory (entry sizes), never the entry
    bodies, so this is cheap and itself allocation-bounded. A non-ZIP buffer
    raises :class:`zipfile.BadZipFile`, which the upload view already treats
    as "not a readable Excel file" — so this can run unconditionally before
    :func:`openpyxl.load_workbook`.
    """
    position = buffer.tell()
    try:
        buffer.seek(0)
        with zipfile.ZipFile(buffer) as archive:
            infos = archive.infolist()
        total_uncompressed = sum(info.file_size for info in infos)
        total_compressed = sum(info.compress_size for info in infos)
    finally:
        buffer.seek(position)

    if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise XlsxTooLargeError(
            f"declared uncompressed size {total_uncompressed} bytes exceeds "
            f"the {MAX_XLSX_UNCOMPRESSED_BYTES}-byte limit"
        )
    if (
        total_compressed > 0
        and total_uncompressed / total_compressed > MAX_XLSX_COMPRESSION_RATIO
    ):
        raise XlsxTooLargeError(
            f"compression ratio {total_uncompressed / total_compressed:.0f}x "
            f"exceeds the {MAX_XLSX_COMPRESSION_RATIO}x limit"
        )


def looks_like_xls(buffer: BinaryIO) -> bool:
    """True if ``buffer`` starts with the OLE2 magic; position is restored.

    Detection is by content, not filename — the production 500 of 2026-07-22
    was an ``.xls`` whose body happened to embed a zip signature, so trusting
    anything but the leading magic bytes misclassifies.
    """
    position = buffer.tell()
    try:
        return buffer.read(len(OLE2_MAGIC)) == OLE2_MAGIC
    finally:
        buffer.seek(position)


def convert_xls_to_xlsx(buffer: BinaryIO) -> io.BytesIO:
    """Convert an in-memory ``.xls`` to an in-memory ``.xlsx``, values only.

    Formatting, formulas, and merged-cell metadata are dropped deliberately —
    parsers read cell values only. Date cells (xlrd ctype ``XL_CELL_DATE``)
    become real ``datetime`` values, matching what openpyxl would give a
    parser for a native ``.xlsx``. Empty cells stay ``None``.

    Raises ``xlrd.XLRDError`` for a corrupt or unsupported OLE2 payload; the
    upload view maps that to its friendly "not a valid export" error.
    """
    source = xlrd.open_workbook(file_contents=buffer.read())
    converted = Workbook(write_only=True)
    for sheet in source.sheets():
        target = converted.create_sheet(title=sheet.name)
        for row_index in range(sheet.nrows):
            values = []
            for cell in sheet.row(row_index):
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    values.append(None)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(
                        xlrd.xldate.xldate_as_datetime(cell.value, source.datemode)
                    )
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    values.append(bool(cell.value))
                elif cell.ctype in (xlrd.XL_CELL_ERROR, xlrd.XL_CELL_BLANK):
                    values.append(None)
                else:
                    values.append(cell.value)
            target.append(values)
    output = io.BytesIO()
    converted.save(output)
    output.seek(0)
    return output
