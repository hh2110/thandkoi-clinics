"""Aggregate a clinic Excel export into de-identified numbers — in memory only.

This is the heart of privacy invariant #1 (CLAUDE.md): the daily export contains
full patient health information, but nothing raw is written to disk or the
database. The export is read from an in-memory stream, tallied into category
counts, and the raw rows are dropped when this function returns.

``ClinicAggregate`` is the only thing that leaves this module. It holds counts,
never names, MRNs, phone numbers, or any other direct identifier — see
``apps.pipeline.ai.PATIENT_IDENTIFYING_COLUMNS`` for the columns that are read
for tallying but never carried forward.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import BinaryIO

from openpyxl import load_workbook

# Columns we read to build counts. Everything not listed here (names, MRNs,
# phone numbers, addresses, free-text notes) is ignored entirely — it is never
# read into the aggregate.
GENDER_COLUMN = "gender"
DIAGNOSIS_COLUMN = "diagnosis"


@dataclass(frozen=True)
class ClinicAggregate:
    """De-identified daily totals. Contains only counts — never patient data.

    Every field here is safe to persist, to send to an AI model, and to publish.
    The frozen dataclass makes the aggregate immutable once computed, so a later
    step can't accidentally fold a raw value back in.
    """

    total_patients: int
    by_gender: dict[str, int] = field(default_factory=dict)
    by_diagnosis: dict[str, int] = field(default_factory=dict)

    def as_dict(self) -> dict[str, object]:
        """Plain, JSON-serialisable representation of the aggregate."""
        return {
            "total_patients": self.total_patients,
            "by_gender": dict(self.by_gender),
            "by_diagnosis": dict(self.by_diagnosis),
        }


def _header_index(header_row: tuple, column_name: str) -> int | None:
    """Find a column by (case-insensitive) header name, or return None."""
    wanted = column_name.strip().lower()
    for index, cell in enumerate(header_row):
        if cell is not None and str(cell).strip().lower() == wanted:
            return index
    return None


def aggregate_export(source: BinaryIO) -> ClinicAggregate:
    """Read an ``.xlsx`` stream and return de-identified counts.

    ``source`` is any binary, seekable stream (an uploaded file, a ``BytesIO``).
    The workbook is opened read-only; no file is written and no database row is
    created. Raw rows exist only as short-lived locals inside this loop and are
    gone once the function returns.
    """
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            return ClinicAggregate(total_patients=0)

        gender_idx = _header_index(header, GENDER_COLUMN)
        diagnosis_idx = _header_index(header, DIAGNOSIS_COLUMN)

        total = 0
        by_gender: dict[str, int] = {}
        by_diagnosis: dict[str, int] = {}

        for row in rows:
            # A row with no cells at all is padding, not a patient.
            if row is None or all(cell is None for cell in row):
                continue
            total += 1

            if gender_idx is not None and gender_idx < len(row):
                value = row[gender_idx]
                if value is not None:
                    key = str(value).strip().lower()
                    by_gender[key] = by_gender.get(key, 0) + 1

            if diagnosis_idx is not None and diagnosis_idx < len(row):
                value = row[diagnosis_idx]
                if value is not None:
                    key = str(value).strip().lower()
                    by_diagnosis[key] = by_diagnosis.get(key, 0) + 1

        # Sort the count maps so the aggregate is deterministic regardless of
        # row order in the source file.
        return ClinicAggregate(
            total_patients=total,
            by_gender=dict(sorted(by_gender.items())),
            by_diagnosis=dict(sorted(by_diagnosis.items())),
        )
    finally:
        workbook.close()
