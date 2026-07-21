"""The first concrete export parser — ``clinic_daily_export_v1``.

**Provisional schema, documented simplification.** The plan (PR #15) treats
the clinic's real export shape as a *sequencing dependency*: a de-identified
or synthetic sample "can be provided before Plan 08 is built", not a design
blocker. No such sample has landed yet, so this parser is written against a
**hand-documented, best-guess column schema** — the columns a clinic EMR
"daily export" commonly carries (patient identity + demographics + visit +
diagnosis + billing), using the same column-name vocabulary already fixed by
``apps.pipeline.ai.PATIENT_IDENTIFYING_COLUMNS`` (Plan 02) for the
identifying columns.

This is a deliberate stand-in, not a guess dressed up as the real thing: the
parser registry's whole design point (one hand-written subclass per format,
no change to the core) means swapping in the *real* parser once the sample
lands is a new subclass + registering it — this file and its tests stay as a
worked example / fallback format, not something that has to be torn up.
Column names below are matched case-insensitively and independent of column
order, so minor real-world header variations (spacing, capitalisation) don't
require a new parser.

Expected columns (case-insensitive, any order):

* ``patient_name``, ``father_name``, ``mrn``, ``phone``, ``address``,
  ``dob`` — direct identifiers. Read only to compute ``age_band`` (from
  ``dob``) and are otherwise dropped. Never attached to a
  :class:`~apps.pipeline.parser_registry.ParsedVisitRow`.
* ``visit_date`` — the clinic-date this row belongs to (required).
* ``department`` — free text, kept as-is (not a direct identifier).
* ``gender`` — mapped via
  :func:`~apps.pipeline.parser_registry.normalise_sex`.
* ``location`` — coarse (village / union council); kept, per the maintainer's
  decision to retain location at this granularity (never the full
  ``address`` column).
* ``diagnosis`` — free text; mapped via
  :func:`~apps.pipeline.parser_registry.diagnosis_category_for`, raw value
  never persisted.
* ``patient_type`` — ``"new"`` or ``"follow-up"`` (case-insensitive);
  anything else maps to unknown (``None``).
* ``payment_type`` — ``"zakat"`` or ``"paid"`` (case-insensitive); anything
  else maps to unknown (``None``).
"""

from __future__ import annotations

import datetime
from typing import BinaryIO

from openpyxl import load_workbook

from apps.pipeline.parser_registry import (
    BaseExportParser,
    ParsedExport,
    ParsedVisitRow,
    ParserRegistry,
    age_band_for,
    diagnosis_category_for,
    header_index,
    normalise_sex,
)

_REQUIRED_COLUMNS = ("visit_date", "gender", "diagnosis")


def _as_date(value) -> datetime.date | None:
    """Coerce an openpyxl cell value to a ``date``, or ``None`` if it can't be."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _as_bool_choice(
    value, *, true_values: set[str], false_values: set[str]
) -> bool | None:
    """Map a free-text cell to True/False/unknown against explicit value sets."""
    text = str(value).strip().lower() if value is not None else ""
    if text in true_values:
        return True
    if text in false_values:
        return False
    return None


class ClinicDailyExportV1Parser(BaseExportParser):
    format_key = "clinic_daily_export_v1"
    label = "Clinic daily export (v1 — provisional schema)"

    def sniff(self, workbook) -> bool:
        sheet = workbook.active
        try:
            header = next(sheet.iter_rows(values_only=True))
        except StopIteration:
            return False
        return all(
            header_index(header, column) is not None for column in _REQUIRED_COLUMNS
        )

    def parse(self, buffer: BinaryIO) -> ParsedExport:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)

            try:
                header = next(rows_iter)
            except StopIteration:
                return ParsedExport(rows=[])

            col = {
                name: header_index(header, name)
                for name in (
                    "visit_date",
                    "department",
                    "gender",
                    "location",
                    "diagnosis",
                    "patient_type",
                    "payment_type",
                    "dob",
                )
            }

            def cell(row, name):
                index = col[name]
                return row[index] if index is not None and index < len(row) else None

            parsed_rows: list[ParsedVisitRow] = []
            for row in rows_iter:
                if row is None or all(value is None for value in row):
                    continue

                visit_date = _as_date(cell(row, "visit_date"))
                if visit_date is None:
                    # A row with no usable visit date can't be attributed to a
                    # clinic-date; skip rather than guess which date it belongs to.
                    continue

                dob = _as_date(cell(row, "dob"))
                age_band = age_band_for(dob=dob, age_years=None, on=visit_date)

                parsed_rows.append(
                    ParsedVisitRow(
                        visit_date=visit_date,
                        department=str(cell(row, "department") or "").strip(),
                        age_band=age_band,
                        sex=normalise_sex(cell(row, "gender")),
                        location=str(cell(row, "location") or "").strip(),
                        diagnosis_category=diagnosis_category_for(
                            cell(row, "diagnosis")
                        ),
                        is_new_patient=_as_bool_choice(
                            cell(row, "patient_type"),
                            true_values={"new"},
                            false_values={
                                "follow-up",
                                "followup",
                                "follow up",
                                "review",
                            },
                        ),
                        is_zakat_beneficiary=_as_bool_choice(
                            cell(row, "payment_type"),
                            true_values={"zakat"},
                            false_values={"paid", "cash", "self-pay", "self pay"},
                        ),
                    )
                )
            return ParsedExport(rows=parsed_rows)
        finally:
            workbook.close()


def register() -> None:
    """Called from ``PipelineConfig.ready()`` so this parser self-registers."""
    ParserRegistry.register(ClinicDailyExportV1Parser())
