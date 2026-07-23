"""Parser for the clinic system's real daily export — ``tkc_daily_activity_v1``.

Grounded in the first real sample (received 2026-07-22, a ``.xls`` converted
in memory by ``apps.pipeline.xls_compat``): sheet ``Patient Report``, two
banner rows (``THE THANDKOI CLINICS — Daily Activity Report`` and
``Period: 08 Jul 2026 to 08 Jul 2026``), a blank row, then a 27-column
header row followed by one row per patient visit. This replaces the
provisional-schema assumption documented in ``parser_clinic_v1`` for real
uploads; that parser is no longer registered as a selectable format
(decision, 2026-07-22 — see its module docstring), though it stays in the
tree as a worked example and as test fixture data.

Layout facts observed on the sample (matched case-insensitively; the header
row is *found*, not assumed at a fixed index, so extra banner lines don't
break parsing):

* There is **no per-row visit date** — every row belongs to the report's
  ``Period:`` line. A daily report has an equal start and end date; a
  multi-day range would make per-row attribution a guess, so it raises
  :class:`~apps.pipeline.parser_registry.ExportParseError` instead.
* ``Sex`` → :func:`~apps.pipeline.parser_registry.normalise_sex`.
* ``Date of Birth`` — text like ``08-Jul-1990`` (often empty); read only as a
  local to compute ``age_band``, then discarded (same rule as
  ``parser_clinic_v1``).
* ``Provisional Diagnosis`` — free text, mapped via
  :func:`~apps.pipeline.parser_registry.diagnosis_category_for`; raw value
  never persisted. Empty on the whole first sample.
* ``Status`` — the payment column: ``zakat`` / ``regular`` →
  ``is_zakat_beneficiary`` True / False; anything else unknown.
* Never read: ``MR #`` (except to check blank/non-blank — see the
  phantom-row note below), ``Patient Name``, ``Father's / Husband's Name``,
  ``Address``, all vitals (BP … Waist). No code path locates their column
  positions except the three identifying headers used to *sniff*, plus
  ``MR #`` itself.
* **The narrative columns are now read** (Plan 11 Track B8/B9, maintainer
  decision 2026-07-23) — ``Presenting Complaints``, ``Investigation``,
  ``Provisional Diagnosis`` (the raw text, kept alongside — not instead of —
  the fixed ``diagnosis_category`` derived from the same column),
  ``Prescribed Medicine``, ``Doctor's Notes`` / ``Nurse's Notes`` /
  ``Dietitian's Notes`` (whichever are present, concatenated into one
  ``clinical_notes`` value), ``Diet & Drug Compliance``, and ``Plan``. This
  reverses the "never read" note this docstring previously carried for these
  columns — see ``apps.pipeline.freetext``'s module docstring for why
  reading raw text from *these specific columns* doesn't violate CLAUDE.md
  invariant #2. **Header text confirmed (2026-07-23)** against the real
  ``TKC july 20th Stat.xls`` sample (header row only — never patient rows):
  every guessed name (``Presenting Complaints``, ``Investigation``,
  ``Provisional Diagnosis``, ``Prescribed Medicine``, ``Doctor's Notes``,
  ``Nurse's Notes``, ``Dietitian's Notes``, ``Diet & Drug Compliance``,
  ``Plan``) matches exactly, straight apostrophes included. ``header_index``
  still degrades harmlessly (field stays blank) if a future export's header
  text ever drifts.
* No department / location / new-vs-follow-up signal exists in this format;
  those fields stay empty/unknown rather than being inferred.

**Phantom continuation rows (bug found 2026-07-22, real 7-patient sample
published as 17).** The clinic system's ``.xls`` writer spills a free-text
column's wrapped content (observed on ``Presenting Complaints`` and
``Prescribed Medicine``) onto extra physical spreadsheet rows when its text
has multiple lines. Those continuation rows carry a value in only that one
free-text column — every other cell, including ``MR #``, is ``None`` — so the
old "is this row entirely blank" check (``all(value is None for value in
row)``) didn't catch them and each one was counted as its own phantom visit.
Fix: a row only counts as a visit if its ``MR #`` cell is non-blank — every
genuine visit row carries one, and no continuation or leftover-formatting row
does. The value itself is still never read into a ``ParsedVisitRow`` or
persisted anywhere, per invariant #1 above; it is inspected only for
blank/non-blank.
"""

from __future__ import annotations

import datetime
import re
from typing import BinaryIO

from openpyxl import load_workbook

from apps.pipeline.parser_registry import (
    BaseExportParser,
    ExportParseError,
    ParsedExport,
    ParsedVisitRow,
    ParserRegistry,
    age_band_for,
    diagnosis_category_for,
    header_index,
    normalise_sex,
)

#: Headers that identify this format's header row (case-insensitive).
_SIGNATURE_COLUMNS = ("mr #", "sex", "provisional diagnosis")

#: ``Period: 08 Jul 2026 to 08 Jul 2026``
_PERIOD_RE = re.compile(r"period:\s*(.+?)\s+to\s+(.+?)\s*$", re.IGNORECASE)

_PERIOD_DATE_FORMAT = "%d %b %Y"
_DOB_FORMAT = "%d-%b-%Y"

#: How many leading rows may precede the header (banners, blanks, and slack).
_HEADER_SEARCH_ROWS = 10


def _find_header(rows: list[tuple]) -> int | None:
    """Index of the first row carrying all three signature headers, else None."""
    for index, row in enumerate(rows[:_HEADER_SEARCH_ROWS]):
        if all(header_index(row, column) is not None for column in _SIGNATURE_COLUMNS):
            return index
    return None


def _period_dates(rows: list[tuple]) -> tuple[datetime.date, datetime.date] | None:
    """The ``Period:`` line's (start, end), scanned from the pre-header rows."""
    for row in rows[:_HEADER_SEARCH_ROWS]:
        for cell in row:
            if not isinstance(cell, str):
                continue
            match = _PERIOD_RE.search(cell)
            if not match:
                continue
            try:
                start = datetime.datetime.strptime(
                    match.group(1).strip(), _PERIOD_DATE_FORMAT
                ).date()
                end = datetime.datetime.strptime(
                    match.group(2).strip(), _PERIOD_DATE_FORMAT
                ).date()
            except ValueError:
                return None
            return start, end
    return None


def _is_blank(value) -> bool:
    """True for ``None`` or a string that's empty once whitespace is trimmed."""
    return value is None or (isinstance(value, str) and value.strip() == "")


def _as_dob(value) -> datetime.date | None:
    """Coerce the text ``Date of Birth`` cell (``08-Jul-1990``) to a date."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return datetime.datetime.strptime(value.strip(), _DOB_FORMAT).date()
        except ValueError:
            return None
    return None


class TkcDailyActivityV1Parser(BaseExportParser):
    format_key = "tkc_daily_activity_v1"
    label = "TKC daily activity report (clinic system .xls)"

    def sniff(self, workbook) -> bool:
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(max_row=_HEADER_SEARCH_ROWS, values_only=True):
            rows.append(row)
        return _find_header(rows) is not None

    def parse(self, buffer: BinaryIO) -> ParsedExport:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        header_at = _find_header(rows)
        if header_at is None:
            raise ExportParseError(
                "Couldn't find the report's column header row — this doesn't "
                "look like a TKC daily activity report."
            )

        period = _period_dates(rows[:header_at])
        if period is None:
            raise ExportParseError(
                "Couldn't read the report's 'Period: ... to ...' line, which "
                "is the only place this format carries the visit date."
            )
        start, end = period
        if start != end:
            raise ExportParseError(
                f"This report covers {start:%d %b %Y} to {end:%d %b %Y}. "
                "Multi-day exports aren't supported — rows can't be "
                "attributed to a single clinic date. Export one day at a time."
            )
        visit_date = start

        header = rows[header_at]
        col = {
            name: header_index(header, name)
            for name in (
                "mr #",
                "sex",
                "date of birth",
                "provisional diagnosis",
                "status",
                # Plan 11 Track B8/B9 free-text columns (2026-07-23) — see
                # the module docstring's caveat on the exact header text for
                # the ones not yet confirmed against a real sample.
                "presenting complaints",
                "investigation",
                "prescribed medicine",
                "doctor's notes",
                "nurse's notes",
                "dietitian's notes",
                "diet & drug compliance",
                "plan",
            )
        }

        def cell(row, name):
            index = col[name]
            return row[index] if index is not None and index < len(row) else None

        def text_cell(row, name) -> str:
            """``cell()``, coerced to a stripped string — "" if blank/missing."""
            return str(cell(row, name) or "").strip()

        parsed_rows: list[ParsedVisitRow] = []
        for row in rows[header_at + 1 :]:
            if row is None:
                continue
            # A genuine visit row always carries its "MR #"; a wrapped-text
            # continuation row (or other leftover blank row) never does — see
            # the module docstring's phantom-row note. This subsumes the old
            # all-cells-None check: a fully blank row also has a blank MR #.
            # Note this also means a free-text value that wraps onto such a
            # continuation row (the exact shape of that bug) is not stitched
            # back on here — only the first line, captured on the genuine
            # visit row itself, is kept.
            if _is_blank(cell(row, "mr #")):
                continue

            dob = _as_dob(cell(row, "date of birth"))
            status = str(cell(row, "status") or "").strip().lower()

            clinical_notes = "; ".join(
                f"{role}: {text}"
                for role, text in (
                    ("Doctor", text_cell(row, "doctor's notes")),
                    ("Nurse", text_cell(row, "nurse's notes")),
                    ("Dietitian", text_cell(row, "dietitian's notes")),
                )
                if text
            )

            parsed_rows.append(
                ParsedVisitRow(
                    visit_date=visit_date,
                    department="",
                    age_band=age_band_for(dob=dob, age_years=None, on=visit_date),
                    sex=normalise_sex(cell(row, "sex")),
                    location="",
                    diagnosis_category=diagnosis_category_for(
                        cell(row, "provisional diagnosis")
                    ),
                    is_new_patient=None,
                    is_zakat_beneficiary=(
                        True
                        if status == "zakat"
                        else False
                        if status == "regular"
                        else None
                    ),
                    presenting_complaints=text_cell(row, "presenting complaints"),
                    investigation=text_cell(row, "investigation"),
                    provisional_diagnosis_text=text_cell(row, "provisional diagnosis"),
                    prescribed_medicine=text_cell(row, "prescribed medicine"),
                    clinical_notes=clinical_notes,
                    diet_and_drug_compliance=text_cell(row, "diet & drug compliance"),
                    plan_notes=text_cell(row, "plan"),
                )
            )
        return ParsedExport(rows=parsed_rows)


def register() -> None:
    """Called from ``PipelineConfig.ready()`` so this parser self-registers."""
    ParserRegistry.register(TkcDailyActivityV1Parser())
